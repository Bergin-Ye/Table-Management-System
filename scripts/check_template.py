import openpyxl

wb = openpyxl.load_workbook("E:/code/System/送货记录模板.xlsx")
ws = wb.active

print(f"Sheet: {ws.title}, Rows: {ws.max_row}, Cols: {ws.max_column}")
print()

# Show header row
headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
print(f"Headers ({len(headers)}): {headers}")
print()

# Find material_code column index (should be "物料编码")
mc_col = None
for i, h in enumerate(headers):
    if h and '物料编码' in str(h):
        mc_col = i + 1
        break

if mc_col is None:
    print("Could not find 物料编码 column")
else:
    print(f"物料编码 is column {mc_col}")
    print()

    # Show rows with empty material_code
    empty_rows = []
    for row in range(2, ws.max_row + 1):
        val = ws.cell(row, mc_col).value
        if val is None or str(val).strip() == '':
            # Show what IS in this row
            row_data = {}
            for c in range(1, len(headers) + 1):
                v = ws.cell(row, c).value
                if v is not None and str(v).strip():
                    row_data[headers[c-1]] = v
            empty_rows.append((row, row_data))

    print(f"Empty material_code rows: {len(empty_rows)}")
    print()
    for row_num, data in empty_rows[:10]:
        print(f"  Row {row_num}: {data}")
    if len(empty_rows) > 10:
        print(f"  ... and {len(empty_rows) - 10} more")
