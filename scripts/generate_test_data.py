#!/usr/bin/env python3
"""生成8个表格的测试数据，每个Excel文件100条记录，中文列名"""
import random
from datetime import date, datetime, timedelta
from decimal import Decimal
from openpyxl import Workbook

ROOT = r"e:\code\System"

# ── 公共数据池 ──
CATEGORIES = ["电控类", "机械类", "气动类", "液压类", "润滑类", "冷却类", "传感器类", "辅料类"]
FACTORIES = ["B1", "B4", "B5", "B6", "B8", "A1", "A3", "C8"]
BRANDS = ["FANUC", "Brother", "TaiKan", "Mitsubishi", "Siemens", "Yaskawa", "Omron", "Schneider"]
MACHINE_MODELS = ["α-D14MiB5", "α-D21MiB5", "α-D14LiB5", "S500X1", "S700X1", "TC-500", "TC-700"]
SHIFTS = ["白班", "夜班"]
UNITS = ["个", "台", "套", "件", "根", "条", "盒", "卷"]
PRODUCT_ATTRS = ["新品", "维修品"]
WARRANTY_STATUS = ["是", "否"]
MACHINE_NOS = [f"D{i:03d}" for i in range(1, 21)] + [f"B{i:02d}" for i in range(1, 9)]
SYSTEM_NAMES = ["CNC系统", "伺服系统", "主轴系统", "刀库系统", "冷却系统", "润滑系统", "排屑系统", "液压系统"]
PART_NAMES = [
    "电源供应器", "I/O板", "伺服放大器", "主轴电机", "刀库电机", "冷却泵", "润滑泵",
    "安全门锁", "限位开关", "接近开关", "电磁阀", "气缸", "油缸", "过滤器", "密封圈",
    "轴承", "丝杠", "导轨", "联轴器", "编码器", "接触器", "继电器", "断路器", "保险丝"
]
MATERIAL_CODES = [f"1529{random.randint(7000,7999)}-{random.randint(10,99):02d}" for _ in range(200)]
MATERIAL_NAMES = [
    "电源供应器", "I/O接口板", "伺服驱动器", "主轴放大器", "冷却风扇", "润滑分配器",
    "安全门开关", "限位传感器", "接近检测器", "电磁换向阀", "气缸组件", "液压油缸",
    "过滤网", "密封垫圈", "滚珠丝杠", "线性导轨", "弹性联轴器", "旋转编码器",
    "交流接触器", "中间继电器", "断路器", "保险管", "整流模块", "变频器"
]
SPEC_MODELS = [
    "A14L-0156-0001#24R", "A20B-2002-0470/11B", "A06B-6050-K060",
    "D4NL-1DFA-BS", "A290-6078-X115", "CF10", "3.2A", "0.75M",
    "500/700型", "16T", "WF300BF", "TEST-001-型", "TEST-002-型"
]

DIAGNOSTICIANS = ["张工", "李工", "王工", "赵工", "陈工", "刘工", "黄工", "周工"]
REPAIR_PERSONS = ["何师傅", "郑师傅", "罗师傅", "梁师傅", "谢师傅", "曾师傅", "吴师傅", "许师傅"]
CONFIRMERS = ["主管-A", "主管-B", "主管-C", "主管-D"]

FAULT_PHENOMENA = ["异响", "振动过大", "温度过高", "无法启动", "精度偏差", "漏油", "漏气", "卡死", "动作迟缓", "报警停机"]
FAULT_DESCS = [
    "运行时有明显金属摩擦异响，疑似轴承磨损",
    "主轴振动超限，加工精度不达标",
    "电机温升过快，超过80℃报警阈值",
    "按下启动按钮无响应，PLC未收到信号",
    "定位偏差0.05mm，超出公差范围",
    "液压管路接头处滴油，密封圈老化",
    "气缸活塞处有明显漏气声",
    "刀库换刀卡住，机械臂不回原位",
    "冷却液出水缓慢，管路堵塞",
    "伺服驱动器报AL-24过载报警"
]

PRICE_TYPES = ["含税", "不含税"]
WARRANTY_PERIODS = ["12个月", "18个月", "24个月", "36个月", "6个月"]

YEAR_MONTHS = ["FY2601", "FY2602", "FY2603", "FY2604", "FY2605", "FY2606", "FY2607"]

def rand_date(start="2026-01-01", end="2026-07-15"):
    d0 = date.fromisoformat(start)
    d1 = date.fromisoformat(end)
    return d0 + timedelta(days=random.randint(0, (d1 - d0).days))

def rand_datetime(start="2026-01-01", end="2026-07-15"):
    d = rand_date(start, end)
    h = random.randint(0, 23)
    m = random.randint(0, 59)
    s = random.randint(0, 59)
    return datetime(d.year, d.month, d.day, h, m, s)

def save(wb, name):
    path = f"{ROOT}\\{name}"
    wb.save(path)
    print(f"  [OK] {name}")

# ═══════════════════════════════════════════
# 1. 送货记录
# ═══════════════════════════════════════════
wb = Workbook()
ws = wb.active
ws.title = "送货记录"
ws.append(["日期", "类别", "物料名称", "规格型号", "物料编码", "物料序列号",
           "数量", "单位", "品牌", "产品属性", "厂房", "出厂单号", "年+月"])
for i in range(1, 101):
    rdate = rand_date()
    ym = f"FY{rdate.strftime('%y%m')}"
    ws.append([
        rdate.strftime("%Y-%m-%d"),
        random.choice(CATEGORIES),
        random.choice(MATERIAL_NAMES),
        random.choice(SPEC_MODELS),
        random.choice(MATERIAL_CODES),
        f"SN{random.randint(10000,99999)}",
        random.randint(1, 50),
        random.choice(UNITS),
        random.choice(BRANDS),
        random.choice(PRODUCT_ATTRS),
        random.choice(FACTORIES),
        f"JS{rdate.strftime('%y%m%d')}{random.randint(100,999)}",
        ym,
    ])
save(wb, "送货记录.xlsx")

# ═══════════════════════════════════════════
# 2. 原始记录
# ═══════════════════════════════════════════
wb = Workbook()
ws = wb.active
ws.title = "原始记录"
ws.append(["年+月", "日期", "班次", "厂房", "序号", "机台号", "诊断人", "维修人",
           "报修时间", "开始时间", "结束时间", "维修工时", "停机工时", "机型",
           "故障现象", "故障描述", "物料编码", "零件名称", "数量", "上机物料",
           "下机物料", "备注", "确认人", "送货记录引用", "上次上机时间", "是否过保"])
for i in range(1, 101):
    rdate = rand_date()
    ym = f"FY{rdate.strftime('%y%m')}"
    req_time = rand_datetime()
    start_time = req_time + timedelta(hours=random.randint(1, 4))
    end_time = start_time + timedelta(hours=random.randint(1, 8))
    rep_h = round((end_time - start_time).total_seconds() / 3600, 1)
    down_h = round(rep_h + random.uniform(0.5, 3), 1)
    last_on = rdate - timedelta(days=random.randint(30, 365))
    ws.append([
        ym,
        rdate.strftime("%Y-%m-%d"),
        random.choice(SHIFTS),
        random.choice(FACTORIES),
        f"XR-{i:04d}",
        random.choice(MACHINE_NOS),
        random.choice(DIAGNOSTICIANS),
        random.choice(REPAIR_PERSONS),
        req_time.strftime("%Y-%m-%d %H:%M:%S"),
        start_time.strftime("%Y-%m-%d %H:%M:%S"),
        end_time.strftime("%Y-%m-%d %H:%M:%S"),
        Decimal(str(rep_h)),
        Decimal(str(down_h)),
        random.choice(MACHINE_MODELS),
        random.choice(FAULT_PHENOMENA),
        random.choice(FAULT_DESCS),
        random.choice(MATERIAL_CODES),
        random.choice(PART_NAMES),
        random.randint(1, 20),
        random.choice(MATERIAL_CODES),
        random.choice(MATERIAL_CODES),
        random.choice(["已修复", "待观察", "更换备件", ""]),
        random.choice(CONFIRMERS),
        f"SHIP-{random.randint(1,99):03d}",
        last_on.strftime("%Y-%m-%d"),
        random.choice(WARRANTY_STATUS),
    ])
save(wb, "原始记录.xlsx")

# ═══════════════════════════════════════════
# 3. 上机物料
# ═══════════════════════════════════════════
wb = Workbook()
ws = wb.active
ws.title = "上机物料"
ws.append(["年+月", "日期", "班次", "厂房", "序号", "机台号", "维修人",
           "报修时间", "开始时间", "结束时间", "维修工时", "停机工时", "机型",
           "故障现象", "故障描述", "物料编码", "零件名称", "数量", "上机物料",
           "下机物料", "备注", "确认人", "送货记录引用", "上次上机时间", "是否过保"])
for i in range(1, 101):
    rdate = rand_date()
    ym = f"FY{rdate.strftime('%y%m')}"
    req_time = rand_datetime()
    start_time = req_time + timedelta(hours=random.randint(1, 4))
    end_time = start_time + timedelta(hours=random.randint(1, 8))
    rep_h = round((end_time - start_time).total_seconds() / 3600, 1)
    down_h = round(rep_h + random.uniform(0.5, 3), 1)
    last_on = rdate - timedelta(days=random.randint(30, 365))
    ws.append([
        ym,
        rdate.strftime("%Y-%m-%d"),
        random.choice(SHIFTS),
        random.choice(FACTORIES),
        f"SM-{i:04d}",
        random.choice(MACHINE_NOS),
        random.choice(REPAIR_PERSONS),
        req_time.strftime("%Y-%m-%d %H:%M:%S"),
        start_time.strftime("%Y-%m-%d %H:%M:%S"),
        end_time.strftime("%Y-%m-%d %H:%M:%S"),
        Decimal(str(rep_h)),
        Decimal(str(down_h)),
        random.choice(MACHINE_MODELS),
        random.choice(FAULT_PHENOMENA),
        random.choice(FAULT_DESCS),
        random.choice(MATERIAL_CODES),
        random.choice(PART_NAMES),
        random.randint(1, 20),
        random.choice(MATERIAL_CODES),
        random.choice(MATERIAL_CODES),
        random.choice(["已修复", "待观察", "需返厂", ""]),
        random.choice(CONFIRMERS),
        f"SHIP-{random.randint(1,99):03d}",
        last_on.strftime("%Y-%m-%d"),
        random.choice(WARRANTY_STATUS),
    ])
save(wb, "上机物料.xlsx")

# ═══════════════════════════════════════════
# 4. 超比统计
# ═══════════════════════════════════════════
wb = Workbook()
ws = wb.active
ws.title = "超比统计"
ws.append(["类别", "物料编码", "系统名称", "零件名称", "单台用量", "比例",
           "含税单价", "机台数", "送货数量", "上机数量", "当月返修",
           "约定比例数量", "超比数量合计", "超比含税金额合计", "统计日期", "年+月"])
for i in range(1, 101):
    rdate = rand_date()
    ym = f"FY{rdate.strftime('%y%m')}"
    unit_usage = Decimal(str(round(random.uniform(1, 10), 2)))
    ratio = Decimal(str(round(random.uniform(0.5, 3.0), 2)))
    unit_price = Decimal(str(round(random.uniform(100, 5000), 2)))
    mc = random.randint(10, 200)
    dq = random.randint(mc, mc * 3)
    mq = random.randint(mc, dq)
    repair = random.randint(0, 20)
    agreed = (unit_usage * ratio * mc).quantize(Decimal("0.01"))
    excess_qty = (Decimal(str(dq)) - agreed).quantize(Decimal("0.01"))
    excess_amt = (excess_qty * unit_price).quantize(Decimal("0.01"))
    ws.append([
        random.choice(CATEGORIES),
        random.choice(MATERIAL_CODES),
        random.choice(SYSTEM_NAMES),
        random.choice(PART_NAMES),
        unit_usage,
        ratio,
        unit_price,
        mc,
        dq,
        mq,
        repair,
        agreed,
        excess_qty if excess_qty > 0 else Decimal("0"),
        excess_amt if excess_amt > 0 else Decimal("0"),
        rdate.strftime("%Y-%m-%d"),
        ym,
    ])
save(wb, "超比统计.xlsx")

# ═══════════════════════════════════════════
# 5. 结算机台数
# ═══════════════════════════════════════════
wb = Workbook()
ws = wb.active
ws.title = "结算机台数"
ws.append(["物料编码", "类别", "零件名称", "单台用量", "比例", "含税单价",
           "质保期", "价格类型", "备注", "机型", "结算机台数"])
for i in range(1, 101):
    ws.append([
        random.choice(MATERIAL_CODES),
        random.choice(CATEGORIES),
        random.choice(PART_NAMES),
        Decimal(str(round(random.uniform(1, 10), 2))),
        Decimal(str(round(random.uniform(0.5, 3.0), 2))),
        Decimal(str(round(random.uniform(100, 5000), 2))),
        random.choice(WARRANTY_PERIODS),
        random.choice(PRICE_TYPES),
        random.choice(["", "标准配置", "选配", "紧急采购"]),
        random.choice(MACHINE_MODELS),
        random.randint(10, 500),
    ])
save(wb, "结算机台数.xlsx")

# ═══════════════════════════════════════════
# 6. 机型明细
# ═══════════════════════════════════════════
wb = Workbook()
ws = wb.active
ws.title = "机型明细"
ws.append(["厂房", "机台号", "机台品牌"])
for i in range(1, 101):
    factory = random.choice(FACTORIES)
    brand = random.choice(BRANDS)
    machine_no = f"{factory}-{brand[:2].upper()}-{random.randint(100,999)}"
    ws.append([factory, machine_no, brand])
save(wb, "机型明细.xlsx")

# ═══════════════════════════════════════════
# 7. 开机数量
# ═══════════════════════════════════════════
wb = Workbook()
ws = wb.active
ws.title = "开机数量"
ws.append(["机型", "数量", "比例(%)", "统计月份", "备注"])
for i in range(1, 101):
    ws.append([
        random.choice(MACHINE_MODELS),
        random.randint(1, 50),
        Decimal(str(round(random.uniform(5, 40), 1))),
        f"2026-{random.randint(1,7):02d}",
        random.choice(["", "正常", "新机到厂", "报废一台", "停用维护"]),
    ])
save(wb, "开机数量.xlsx")

# ═══════════════════════════════════════════
# 8. 物料表
# ═══════════════════════════════════════════
wb = Workbook()
ws = wb.active
ws.title = "物料表"
ws.append(["类别", "物料名称", "规格型号", "物料编码"])
for i in range(1, 101):
    ws.append([
        random.choice(CATEGORIES),
        random.choice(MATERIAL_NAMES),
        random.choice(SPEC_MODELS),
        random.choice(MATERIAL_CODES),
    ])
save(wb, "物料表.xlsx")

print("\n[OK] All 8 Excel files generated!")
